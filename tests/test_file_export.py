"""Test plugin file-export — provider xuất CSV/Excel/Word + generalize artifact emit."""

import base64
import io
import json
import shutil

import pytest

from app.tools.file_export import FileExportProvider, _artifact_dir, artifact_b64

_CONV = "test-file-export-conv"


def _cleanup():
    d = _artifact_dir(_CONV)
    if d.exists():
        shutil.rmtree(d.parent if d.name == _CONV else d, ignore_errors=True)


class TestFileExportProvider:
    def setup_method(self):
        self.p = FileExportProvider()
        _cleanup()

    def teardown_method(self):
        _cleanup()

    def test_server_meta(self):
        assert self.p.server_name == "file-export"
        assert self.p.is_mock is False  # tool THẬT

    def test_tools_listed(self):
        names = {t.name for t in self.p.list_tools()}
        assert names == {"export_csv", "export_xlsx", "export_docx"}

    def test_export_csv(self):
        out = self.p.call("export_csv", {
            "_conversation_id": _CONV,
            "filename": "danh-sach",
            "columns": ["Tên", "Số tiền"],
            "rows": [["Nguyễn Văn A", 100000], ["Trần Thị B", 200000]],
        })
        meta = json.loads(out)
        assert meta["artifact"] is True
        assert meta["filename"] == "danh-sach.csv"
        f = _artifact_dir(_CONV) / "danh-sach.csv"
        assert f.is_file()
        text = f.read_text(encoding="utf-8-sig")
        assert "Tên,Số tiền" in text
        assert "Nguyễn Văn A" in text

    def test_export_xlsx(self):
        out = self.p.call("export_xlsx", {
            "_conversation_id": _CONV,
            "columns": ["A", "B"],
            "rows": [[1, 2], [3, 4]],
            "sheet_name": "Báo cáo",
        })
        meta = json.loads(out)
        assert meta["filename"] == "export.xlsx"  # filename trống → fallback
        f = _artifact_dir(_CONV) / "export.xlsx"
        assert f.is_file()
        from openpyxl import load_workbook
        wb = load_workbook(f)
        ws = wb.active
        assert [c.value for c in ws[1]] == ["A", "B"]
        assert [c.value for c in ws[2]] == [1, 2]

    def test_export_docx(self):
        out = self.p.call("export_docx", {
            "_conversation_id": _CONV,
            "title": "Báo cáo tháng",
            "sections": [
                {"heading": "Tổng quan", "paragraphs": ["Đoạn 1.", "Đoạn 2."]},
                {"paragraphs": ["Không có heading."]},
            ],
        })
        meta = json.loads(out)
        assert meta["filename"] == "bao-cao-thang.docx" or meta["filename"].endswith(".docx")
        f = _artifact_dir(_CONV) / meta["filename"]
        assert f.is_file()
        from docx import Document
        doc = Document(f)
        texts = [p.text for p in doc.paragraphs]
        assert "Báo cáo tháng" in texts
        assert "Đoạn 1." in texts

    def test_rows_limit(self):
        with pytest.raises(ValueError):
            self.p.call("export_csv", {
                "_conversation_id": _CONV,
                "columns": ["x"],
                "rows": [[i] for i in range(501)],
            })

    def test_columns_required(self):
        with pytest.raises(ValueError):
            self.p.call("export_csv", {"_conversation_id": _CONV, "columns": [], "rows": []})

    def test_filename_path_traversal_blocked(self):
        out = self.p.call("export_csv", {
            "_conversation_id": _CONV,
            "filename": "../../../etc/evil",
            "columns": ["a"],
            "rows": [["1"]],
        })
        meta = json.loads(out)
        # Base name an toàn, KHÔNG escape khỏi thư mục artifact.
        assert "/" not in meta["filename"]
        assert (_artifact_dir(_CONV) / meta["filename"]).is_file()

    def test_artifact_b64_roundtrip(self):
        self.p.call("export_csv", {
            "_conversation_id": _CONV, "filename": "x", "columns": ["a"], "rows": [["1"]],
        })
        b64 = artifact_b64(_CONV, "x.csv")
        assert b64 is not None
        assert b64decode_ok(b64)

    def test_unknown_tool(self):
        with pytest.raises(ValueError):
            self.p.call("export_pdf", {"_conversation_id": _CONV})


def b64decode_ok(b64: str) -> bool:
    try:
        base64.b64decode(b64)
        return True
    except Exception:
        return False


class TestArtifactEmitGeneralization:
    """ChatEngine._extract_artifact_meta nhận diện cả file-export lẫn Upia (tương thích ngược)."""

    def test_extract_file_export(self):
        from app.core.chat_engine import _extract_artifact_meta
        from app.llm.base import ToolCallEvent, ToolResult

        ev = ToolCallEvent(
            name="file-export__export_csv",
            input={},
            result=ToolResult(content=json.dumps({"artifact": True, "filename": "x.csv", "size_kb": 1.2})),
        )
        meta = _extract_artifact_meta(ev)
        assert meta == {"filename": "x.csv", "size_kb": 1.2}

    def test_extract_upia_backcompat(self):
        from app.core.chat_engine import _extract_artifact_meta
        from app.llm.base import ToolCallEvent, ToolResult

        ev = ToolCallEvent(
            name="partner-integration__package_project",
            input={},
            result=ToolResult(content=json.dumps({"packaged": True, "zip_name": "p.zip", "size_kb": 5.0})),
        )
        meta = _extract_artifact_meta(ev)
        assert meta == {"filename": "p.zip", "size_kb": 5.0}

    def test_extract_non_artifact(self):
        from app.core.chat_engine import _extract_artifact_meta
        from app.llm.base import ToolCallEvent, ToolResult

        ev = ToolCallEvent(name="system__get_current_date", input={}, result=ToolResult(content="2026-06-15"))
        assert _extract_artifact_meta(ev) is None

    def test_extract_error_result(self):
        from app.core.chat_engine import _extract_artifact_meta
        from app.llm.base import ToolCallEvent, ToolResult

        ev = ToolCallEvent(
            name="file-export__export_csv",
            input={},
            result=ToolResult(content=json.dumps({"artifact": True, "filename": "x.csv"}), is_error=True),
        )
        assert _extract_artifact_meta(ev) is None
