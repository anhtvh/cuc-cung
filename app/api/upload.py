"""POST /upload — trích nội dung file để inject vào chat (text/image)."""

import base64
import logging
from io import BytesIO

from fastapi import APIRouter, HTTPException, UploadFile

log = logging.getLogger(__name__)
router = APIRouter(tags=["upload"])

MAX_BYTES = 5 * 1024 * 1024  # 5 MB

_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"
_JPEG_MAGIC = b"\xff\xd8"


def _detect_image_type(raw: bytes) -> str | None:
    if raw[:8] == _PNG_MAGIC:
        return "image/png"
    if raw[:2] == _JPEG_MAGIC:
        return "image/jpeg"
    return None


@router.post("/upload")
async def upload_file(file: UploadFile):
    raw = await file.read()
    if not raw:
        raise HTTPException(422, "File rỗng — không có nội dung để xử lý")
    if len(raw) > MAX_BYTES:
        raise HTTPException(413, "File quá lớn (tối đa 5 MB)")

    name = (file.filename or "").lower()
    ct = (file.content_type or "").split(";")[0].strip()

    if name.endswith((".txt", ".md")) or ct in ("text/plain", "text/markdown"):
        return {"filename": file.filename, "content_type": "text",
                "text": raw.decode("utf-8", errors="replace")}

    if name.endswith(".pdf") or ct == "application/pdf":
        return {"filename": file.filename, "content_type": "text", "text": _extract_pdf(raw)}

    if name.endswith(".docx") or ct == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return {"filename": file.filename, "content_type": "text", "text": _extract_docx(raw)}

    if name.endswith(".csv") or ct in ("text/csv", "application/csv"):
        return {"filename": file.filename, "content_type": "text", "text": _extract_csv(raw)}

    if name.endswith((".xlsx", ".xlsm")) or ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return {"filename": file.filename, "content_type": "text", "text": _extract_excel(raw)}

    if name.endswith((".png", ".jpg", ".jpeg")) or ct in ("image/png", "image/jpeg"):
        # I-07: detect thật bằng magic bytes thay vì tin extension/content-type client gửi
        media_type = _detect_image_type(raw) or ("image/png" if name.endswith(".png") else "image/jpeg")
        return {"filename": file.filename, "content_type": "image",
                "base64": base64.b64encode(raw).decode(), "media_type": media_type}

    raise HTTPException(415, f"Định dạng không hỗ trợ: {file.filename}")


def _extract_pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(raw))
        pages = [page.extract_text() or "" for page in reader.pages]
        text = "\n\n".join(p for p in pages if p.strip())
        if not text:
            raise HTTPException(422, "PDF không có text layer (ảnh scan — thử upload ảnh trực tiếp)")
        return text
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(422, f"Không đọc được PDF: {e}") from e


def _extract_docx(raw: bytes) -> str:
    try:
        import docx
        doc = docx.Document(BytesIO(raw))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as e:
        raise HTTPException(422, f"Không đọc được DOCX: {e}") from e


# Giới hạn text trích từ bảng để tránh blow-up bộ nhớ/context (5MB xlsx có thể bung rất lớn).
# File lớn vẫn đi qua RAG ingest (ngưỡng ~8000 ký tự) ở chat_engine nên cắt ở đây an toàn.
_TABLE_MAX_CHARS = 200_000


def _extract_csv(raw: bytes) -> str:
    """CSV → text. Decode bền (thử utf-8-sig rồi fallback), giữ nguyên dạng phân cách."""
    import csv
    from io import StringIO

    # utf-8-sig để nuốt BOM của Excel-export; lỗi ký tự thì replace (không bỏ lỗi im lặng).
    for enc in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    # Chuẩn hoá về dạng "cell | cell" cho model dễ đọc + thống nhất với Excel.
    try:
        reader = csv.reader(StringIO(text))
        lines = [" | ".join(cell.strip() for cell in row) for row in reader if any(c.strip() for c in row)]
        out = "\n".join(lines)
    except Exception:
        out = text  # parse hỏng → trả raw, vẫn hữu ích hơn là fail
    if not out.strip():
        raise HTTPException(422, "CSV rỗng — không có dữ liệu để xử lý")
    return out[:_TABLE_MAX_CHARS]


def _extract_excel(raw: bytes) -> str:
    """Excel .xlsx/.xlsm → text. Mỗi sheet một khối, mỗi dòng "cell | cell"."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover - openpyxl là dependency bắt buộc
        raise HTTPException(500, "Thiếu thư viện openpyxl để đọc Excel") from e
    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(422, f"Không đọc được Excel: {e}") from e

    parts: list[str] = []
    total = 0
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in row]
            if not any(cells):
                continue
            line = " | ".join(cells).rstrip(" |")
            rows.append(line)
            total += len(line) + 1
            if total >= _TABLE_MAX_CHARS:
                break
        if rows:
            parts.append(f"# Sheet: {ws.title}\n" + "\n".join(rows))
        if total >= _TABLE_MAX_CHARS:
            parts.append("… (đã cắt bớt vì file quá lớn)")
            break
    wb.close()
    out = "\n\n".join(parts)
    if not out.strip():
        raise HTTPException(422, "Excel rỗng — không có dữ liệu để xử lý")
    return out
